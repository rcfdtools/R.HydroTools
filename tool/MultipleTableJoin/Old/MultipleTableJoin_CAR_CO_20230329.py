# -*- coding: UTF-8 -*-

import openpyxl

def remove(sheet, row):
    # iterate the row object
    for cell in row:
          # check the value of each cell in
        # the row, if any of the value is not
        # None return without removing the row
        if cell.value != None:
              return
    # get the row number from the first cell
    # and remove the row
    sheet.delete_rows(row[0].row, 1)

# General parameters
path = 'W:/Descargas/'
excel_file = '5de94fdc2fabb.xlsx'
head_rows = 15  # Header rows to delete
cols_del = [3,5,7,9,11,13,15,17,19,21,23,25]

from openpyxl import Workbook, load_workbook
book = load_workbook(path+excel_file)
catalog = book['Catalogo']
book.remove(catalog)
sheets = book.sheetnames
'''
sheet = book.active
print(book.worksheets)
print(book.sheetnames)
print(sheet['D8'].value)
'''
# Clean headers and intermediate columns
for i in sheets:
    print('Processing: %s' %i)
    sheet = book[i]
    print('\tTotal rows: %d' % sheet.max_row)
    for merge in list(sheet.merged_cells):
        print('\tUnmerging %s' % merge)
        sheet.unmerge_cells(range_string=str(merge))
    #sheet.delete_rows(idx=1)  # Delete an specific row
    sheet.delete_rows(1,head_rows)
    print('\tDeleting headers...')
    print('\tDeleting white columns...')
    for j in range(3,15):  # Delete intermediate columns
        sheet.delete_cols(j)
    print('\tRows: %d' % len(sheet['A'] > 0))
    current_row = 1
    print('\tTotal data rows: %d' % sheet.max_row)
    for each_row in sheet.iter_rows():
        sheet.cell(row=current_row, column=2).value = i
        current_row += 1
'''        
# Include date and station code
for i in sheets:
    sheet = book[i]
    for merge in list(sheet.merged_cells):
        print('Unmerging %s %s' %(sheet, merge))
        sheet.unmerge_cells(range_string=str(merge))
    current_row = 1
    for each_row in sheet.iter_rows():
        sheet.cell(row=current_row, column=2).value=i
        current_row+=1
'''

book.save(path+'Clean_'+excel_file)


# Refs
# https://blog.aspose.com/cells/insert-and-delete-rows-and-columns-in-excel-using-python/
# https://www.youtube.com/watch?v=718edSNvKLA
# https://stackoverflow.com/questions/23527887/getting-sheet-names-from-openpyxl
# https://www.geeksforgeeks.org/how-to-delete-one-or-more-rows-in-excel-using-openpyxl/
# https://www.codespeedy.com/delete-rows-of-a-sheet-using-openpyxl/
# https://stackoverflow.com/questions/69891944/unmerge-every-cell-in-an-excel-worksheet-using-openpyxl
# https://stackoverflow.com/questions/40166714/replace-specific-values-in-openpyxl