# -*- coding: UTF-8 -*-
# Name: MultipleTableJoin_CAR_CO.py
# Description: this script create a dataset file from monthly CAR - Cundinamarca - Colombia Excel records
# Source dataset: https://www.car.gov.co/vercontenido/2524
# Disclaimer: this script version do not create the records marks file and the wind directions file
# Repository: https://github.com/rcfdtools/R.GISPython/tree/main/MultipleTableJoin
# License: https://github.com/rcfdtools/R.GISPython/wiki/License
# Requirements: Python 3+, Pandas, openpyxl


# Libraries
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
import glob


# General parameters
input_path = 'D:/R.GISPython/MultipleTableJoin/CAR_CO/Source/'
output_path = 'D:/R.GISPython/MultipleTableJoin/CAR_CO/Joined/'
parameter_value = 'Q_MEDIA_M'  # More info https://github.com/rcfdtools/R.GISPython/tree/main/MultipleTableJoin/CAR_CO
excel_file = '5bae92275ec7e.xlsx'
clean_file = 'Clean_'+excel_file
pivot_file = 'Pivot_'+parameter_value+'.csv'
unpivot_file = 'Unpivot_'+parameter_value+'.csv'
joined_file = 'CAR_Joined.csv'
head_rows = 15  # Header rows to delete
index_name = 'Id'
parameter_name = 'Label'
station_code = 'Code'
year_name = 'Year'
month_name = 'Month'
day_name = 'Day'
date_name = 'Date'
value_name = 'Value'
run_clean = True  # Execute the clean Excel process. Set False if you are going to join all the files.
run_pivot_dataset = True  # Execute the pivot dataset creation. Set False if you are going to join all the files.
run_unpivot_dataset = True  # Execute the unpivot dataset creation. Set False if you are going to join all the files.
year_to_integer = False  # Convert Year values to integer.
join_unpivot_files = True  # Set to True if you had all the files processed with unpivot tables and you require one joined file.


# Load workbook and delete catalog and not station sheets
book = load_workbook(input_path+excel_file, data_only=True)  # use data_only for read values and not formulas
delete_sheets = ['Catalogo', 'CATALOGO', 'Resumen', 'RESUMEN', 'Hoja1', 'Hoja2', 'Hoja3', 'Sheet1', 'Sheet2', 'Sheet3']
sheets = book.sheetnames
for j in sheets:
    for i in delete_sheets:
        if j == i:
            book.remove(book[i])
sheets = book.sheetnames


# Clean headers and intermediate columns
if run_clean:
    print('\nRunning the cleaning process\n'+ 40 * '-')
    for i in sheets:
        print('Processing: %s' %i)
        sheet = book[i]
        print('\tTotal rows: %d' % sheet.max_row)
        for merge in list(sheet.merged_cells):
            print('\tUnmerging %s' % merge)
            sheet.unmerge_cells(range_string=str(merge))
        sheet.delete_rows(1,head_rows)
        print('\tDeleting headers')
        print('\tDeleting white columns')
        sheet.delete_cols(2)
        for j in range(2,15):  # Delete intermediate columns
            sheet.delete_cols(j)
        print('\tRows: %d' % len(sheet['A']))
    book.save(output_path+clean_file)


# Pivot dataset creation
if run_pivot_dataset:
    print('\n\nRunning the pivot dataframe creation\n'+ 40 * '-')
    xl = pd.ExcelFile(output_path+clean_file)
    sheets = xl.sheet_names
    df = pd.DataFrame()
    for i in sheets:
        print('Processing: %s' % i)
        df1 = pd.read_excel(output_path+clean_file, sheet_name=i, header=None, names=[year_name, '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12'])
        df1[station_code] = i
        df1[parameter_name] = parameter_value
        df = pd.concat([df, df1])
    if year_to_integer:
        df = df.astype({year_name : int})
    print(df)
    df.to_csv(output_path+pivot_file, encoding='utf-8', index=False)


# Unpivot dataset creation
if run_unpivot_dataset:
    print('\n\nRunning the unpivot dataframe creation\n'+ 40 * '-')
    df = pd.read_csv(output_path+pivot_file, low_memory=False)
    print(df.info(), '\n')
    print(df.dtypes, '\n')
    print(df, '\n')
    df_unpivot = pd.melt(df, id_vars=(year_name, station_code, parameter_name), var_name=month_name, value_name=value_name, ignore_index = True)
    df_unpivot[day_name] = '01'
    df_unpivot[date_name] = pd.to_datetime(df_unpivot[[year_name, month_name, day_name]])
    df_unpivot.index.name = index_name
    df_unpivot = df_unpivot[[parameter_name, station_code, year_name, month_name, day_name, date_name, value_name]]
    print(df_unpivot, '\n')
    df_unpivot.to_csv(output_path+unpivot_file)


# Join all the unpivot files
if join_unpivot_files:
    print('\nJoining the unpivot files\n' + 40 * '-')
    csv_files = glob.glob(output_path + 'Unpivot_*.csv')
    print('Joining: %s' %csv_files)
    df = pd.concat(map(pd.read_csv, csv_files), ignore_index=True)
    df.to_csv(output_path + joined_file, encoding='utf-8', index=False)
    print(df)


print('\nClean file: %s' % output_path + clean_file,
          '\nPivot file: %s' % output_path + pivot_file,
          '\nUnpivot file: %s' % output_path + unpivot_file,
          '\nJoined file: %s' % output_path + joined_file)

